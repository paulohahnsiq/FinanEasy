from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from models.conta import Conta

ENTRADAS = ("deposito", "transferencia_recebida")


def gerar_extrato_pdf(conta: Conta, caminho: str = None) -> str:
    """Gera um extrato em PDF para a conta informada e retorna o caminho do arquivo."""
    caminho = caminho or f"extrato_conta_{conta.numero}.pdf"

    pdf = FPDF()
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "FinanEasy - Extrato Bancário", ln=1, align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 8, f"Conta: {conta.numero}", ln=1)
    pdf.cell(0, 8, f"Cliente: {conta.cliente.nome}", ln=1)
    pdf.cell(0, 8, f"CPF: {conta.cliente.cpf}", ln=1)
    pdf.cell(0, 8, f"Saldo atual: R$ {conta.saldo:.2f}", ln=1)
    pdf.ln(6)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(40, 8, "Data", border=1)
    pdf.cell(50, 8, "Tipo", border=1)
    pdf.cell(30, 8, "Valor (R$)", border=1)
    pdf.cell(70, 8, "Descrição", border=1, ln=1)

    pdf.set_font("Helvetica", "", 10)
    for t in conta.transacoes:
        sinal = "+" if t.tipo in ENTRADAS else "-"
        pdf.cell(40, 8, t.data, border=1)
        pdf.cell(50, 8, t.tipo.replace("_", " ").title(), border=1)
        pdf.cell(30, 8, f"{sinal}{t.valor:.2f}", border=1)
        pdf.cell(70, 8, t.descricao[:40], border=1, ln=1)

    pdf.output(caminho)
    return caminho


def gerar_extrato_excel(conta: Conta, caminho: str = None) -> str:
    """Gera um extrato em Excel (.xlsx) para a conta informada e retorna o caminho do arquivo."""
    caminho = caminho or f"extrato_conta_{conta.numero}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Extrato"

    ws["A1"] = "FinanEasy - Extrato Bancário"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"], ws["B3"] = "Conta:", conta.numero
    ws["A4"], ws["B4"] = "Cliente:", conta.cliente.nome
    ws["A5"], ws["B5"] = "CPF:", conta.cliente.cpf
    ws["A6"], ws["B6"] = "Saldo atual:", f"R$ {conta.saldo:.2f}"

    linha_cabecalho = 8
    for col, titulo in enumerate(["Data", "Tipo", "Valor (R$)", "Descrição"], start=1):
        celula = ws.cell(row=linha_cabecalho, column=col, value=titulo)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    linha = linha_cabecalho + 1
    for t in conta.transacoes:
        sinal = 1 if t.tipo in ENTRADAS else -1
        ws.cell(row=linha, column=1, value=t.data)
        ws.cell(row=linha, column=2, value=t.tipo.replace("_", " ").title())
        ws.cell(row=linha, column=3, value=sinal * t.valor)
        ws.cell(row=linha, column=4, value=t.descricao)
        linha += 1

    for i, largura in enumerate([20, 25, 15, 35], start=1):
        ws.column_dimensions[chr(64 + i)].width = largura

    wb.save(caminho)
    return caminho
